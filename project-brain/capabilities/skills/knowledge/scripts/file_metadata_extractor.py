# DEPRECATED: This module is superseded by the document-extractor skill (ADR-518).
# Use `extract-document` MCP tool instead. This file will be removed in a future cleanup.
"""Extract file metadata as JSON for Tier 3 RAG indexing (ADR-085).

Usage:
    python3 file_metadata_extractor.py /path/to/file.pdf

Outputs JSON with at minimum: name, type, size_bytes.
Type-specific fields are added when possible (PDF page count, etc.).
"""

import argparse
import json
import os
import subprocess
import sys


def _extract_pdf(path: str) -> dict:
    """Extract PDF-specific metadata."""
    try:
        from PyPDF2 import PdfReader  # type: ignore

        reader = PdfReader(path)
        return {"page_count": len(reader.pages)}
    except Exception:
        pass

    # macOS fallback via mdls
    try:
        result = subprocess.run(
            ["mdls", "-name", "kMDItemNumberOfPages", path],
            capture_output=True,
            text=True,
            timeout=5,
        )
        if result.returncode == 0:
            # Output looks like: kMDItemNumberOfPages = 3
            value = result.stdout.strip().split("=")[-1].strip()
            if value and value != "(null)":
                return {"page_count": int(value)}
    except Exception:
        pass

    return {}


def _extract_excel(path: str) -> dict:
    """Extract Excel-specific metadata."""
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return {"sheet_names": names}
    except Exception:
        return {}


def _extract_csv(path: str) -> dict:
    """Extract CSV-specific metadata (row count excluding header)."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            line_count = sum(1 for _ in f)
        # Subtract 1 for header row; clamp to 0
        return {"row_count": max(0, line_count - 1)}
    except Exception:
        return {}


def _extract_text(path: str) -> dict:
    """Extract text/markdown metadata."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        lines = content.splitlines()
        words = content.split()
        return {"word_count": len(words), "line_count": len(lines)}
    except Exception:
        return {}


def _extract_json(path: str) -> dict:
    """Extract JSON metadata (top-level keys)."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return {"keys": list(data.keys())}
        return {}
    except Exception:
        return {}


def _extract_yaml(path: str) -> dict:
    """Extract YAML metadata (top-level keys)."""
    try:
        import yaml  # type: ignore

        with open(path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        if isinstance(data, dict):
            return {"keys": list(data.keys())}
        return {}
    except Exception:
        return {}


# Map extensions to extraction functions
_EXTRACTORS = {
    ".pdf": _extract_pdf,
    ".xlsx": _extract_excel,
    ".xls": _extract_excel,
    ".csv": _extract_csv,
    ".md": _extract_text,
    ".txt": _extract_text,
    ".json": _extract_json,
    ".yaml": _extract_yaml,
    ".yml": _extract_yaml,
}


def extract_metadata(path: str) -> dict:
    """Extract metadata for a single file. Always returns name, type, size_bytes."""
    _, ext = os.path.splitext(path)
    ext_lower = ext.lower()
    file_type = ext_lower.lstrip(".") if ext_lower else "unknown"

    metadata = {
        "name": os.path.basename(path),
        "type": file_type,
        "size_bytes": os.path.getsize(path),
    }

    extractor = _EXTRACTORS.get(ext_lower)
    if extractor:
        extra = extractor(path)
        if extra:
            metadata["extracted"] = extra

    return metadata


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract file metadata as JSON for Tier 3 RAG indexing.")
    parser.add_argument("path", help="Path to the file to extract metadata from")
    args = parser.parse_args()

    path = os.path.abspath(args.path)
    if not os.path.isfile(path):
        print(json.dumps({"error": f"File not found: {path}"}), file=sys.stderr)
        sys.exit(1)

    metadata = extract_metadata(path)
    print(json.dumps(metadata, indent=2))


if __name__ == "__main__":
    main()
