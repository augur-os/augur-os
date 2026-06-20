#!/usr/bin/env python3
"""Excel/CSV reader CLI for the Hub Data Bridge (ADR-086).

Standalone CLI tool callable via subprocess from TypeScript API routes.
Reads specific cell values from Excel files or column values from CSVs
and returns formatted JSON.

Usage:
    # Extract specific cells from an Excel file
    python3 excel_reader.py --file balance-sheet.xlsx \\
        --extractions '[{"sheet":"Summary","cell":"B42","label":"Total Assets"}]'

    # Extract column values from a CSV file
    python3 excel_reader.py --file investments.csv \\
        --columns '["Total","Value"]'

    # Extract with format hints
    python3 excel_reader.py --file balance-sheet.xlsx \\
        --extractions '[{"sheet":"Summary","cell":"B42","label":"Total","format":"currency"}]'

Output (JSON array):
    [{"label":"Total Assets","value":125000,"formatted":"$125,000.00"}]
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

# openpyxl is an optional dependency
try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    HAS_OPENPYXL = False


def format_value(value: Any, fmt: str = "auto") -> str:
    """Format a value for display.

    Args:
        value: The raw value to format.
        fmt: Format hint: "currency", "percent", "number", "auto".

    Returns:
        Human-readable formatted string.
    """
    if value is None:
        return ""

    if fmt == "currency":
        try:
            num = float(value)
            if num < 0:
                return f"-${abs(num):,.2f}"
            return f"${num:,.2f}"
        except (ValueError, TypeError):
            return str(value)

    if fmt == "percent":
        try:
            num = float(value)
            return f"{num:.1%}" if abs(num) <= 1 else f"{num:.1f}%"
        except (ValueError, TypeError):
            return str(value)

    if fmt == "number":
        try:
            num = float(value)
            if num == int(num):
                return f"{int(num):,}"
            return f"{num:,.2f}"
        except (ValueError, TypeError):
            return str(value)

    # Auto-detect format
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != int(value):
            return f"{value:,.2f}"
        return f"{int(value):,}"

    return str(value)


def read_excel_extractions(file_path: Path, extractions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Read specific cell values from an Excel file.

    Args:
        file_path: Path to the .xlsx file.
        extractions: List of extraction specs, each with:
            - sheet: Sheet name
            - cell: Cell reference (e.g., "B42")
            - label: Display label
            - format: Optional format hint

    Returns:
        List of result dicts with label, value, formatted.
    """
    if not HAS_OPENPYXL:
        return [
            {
                "label": ex.get("label", ""),
                "value": None,
                "formatted": "",
                "error": "openpyxl not installed",
            }
            for ex in extractions
        ]

    results: list[dict[str, Any]] = []

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as e:
        return [
            {
                "label": ex.get("label", ""),
                "value": None,
                "formatted": "",
                "error": str(e),
            }
            for ex in extractions
        ]

    for ex in extractions:
        sheet_name = ex.get("sheet", "")
        cell_ref = ex.get("cell", "")
        label = ex.get("label", cell_ref)
        fmt = ex.get("format", "auto")

        result: dict[str, Any] = {"label": label, "value": None, "formatted": ""}

        if sheet_name not in wb.sheetnames:
            result["error"] = f"Sheet not found: {sheet_name}"
            results.append(result)
            continue

        ws = wb[sheet_name]
        try:
            cell = ws[cell_ref]
            raw_value = cell.value
            result["value"] = raw_value
            result["formatted"] = format_value(raw_value, fmt)
        except Exception as e:
            result["error"] = str(e)

        results.append(result)

    wb.close()
    return results


def read_csv_columns(file_path: Path, columns: list[str]) -> list[dict[str, Any]]:
    """Read column aggregates from a CSV file.

    For each requested column, returns the last non-empty value (which
    is typically a total row) and the column's values.

    Args:
        file_path: Path to the .csv file.
        columns: List of column header names to extract.

    Returns:
        List of result dicts with label, value, formatted.
    """
    results: list[dict[str, Any]] = []

    try:
        with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None:
                return [{"label": col, "value": None, "formatted": "", "error": "No headers"} for col in columns]

            # Read all rows, keep track of last non-empty value per column
            last_values: dict[str, Any] = {}
            row_count = 0
            for row in reader:
                row_count += 1
                for col in columns:
                    val = row.get(col, "")
                    if val and str(val).strip():
                        last_values[col] = val.strip()

            for col in columns:
                result: dict[str, Any] = {"label": col}
                raw = last_values.get(col)
                if raw is not None:
                    # Try to parse as number
                    try:
                        cleaned = raw.replace(",", "").replace("$", "").replace("%", "")
                        num = float(cleaned)
                        result["value"] = num
                        if "$" in raw:
                            result["formatted"] = format_value(num, "currency")
                        elif "%" in raw:
                            result["formatted"] = format_value(num, "percent")
                        else:
                            result["formatted"] = format_value(num, "number")
                    except ValueError:
                        result["value"] = raw
                        result["formatted"] = raw
                else:
                    result["value"] = None
                    result["formatted"] = ""
                    result["error"] = f"Column not found or empty: {col}"

                results.append(result)

    except Exception as e:
        return [{"label": col, "value": None, "formatted": "", "error": str(e)} for col in columns]

    return results


def main() -> None:
    """CLI entry point."""
    parser = argparse.ArgumentParser(description="Read values from Excel/CSV files and return formatted JSON.")
    parser.add_argument(
        "--file",
        required=True,
        help="Path to Excel (.xlsx) or CSV (.csv) file",
    )
    parser.add_argument(
        "--extractions",
        help='JSON array of extraction specs: [{"sheet":"S","cell":"B2","label":"L","format":"currency"}]',
    )
    parser.add_argument(
        "--columns",
        help='JSON array of column names for CSV: ["Total","Value"]',
    )

    args = parser.parse_args()
    file_path = Path(args.file).expanduser().resolve()

    if not file_path.exists():
        json.dump(
            [{"error": f"File not found: {file_path}"}],
            sys.stdout,
            indent=2,
        )
        sys.exit(1)

    suffix = file_path.suffix.lower()

    if args.extractions:
        try:
            extractions = json.loads(args.extractions)
        except json.JSONDecodeError as e:
            json.dump([{"error": f"Invalid --extractions JSON: {e}"}], sys.stdout, indent=2)
            sys.exit(1)
        results = read_excel_extractions(file_path, extractions)

    elif args.columns:
        try:
            columns = json.loads(args.columns)
        except json.JSONDecodeError as e:
            json.dump([{"error": f"Invalid --columns JSON: {e}"}], sys.stdout, indent=2)
            sys.exit(1)

        if suffix == ".csv":
            results = read_csv_columns(file_path, columns)
        elif suffix in (".xlsx", ".xls"):
            # For Excel, treat columns as "read last value from each column"
            # This is a convenience mode — extractions is more precise
            results = [
                {
                    "label": col,
                    "value": None,
                    "formatted": "",
                    "error": "Use --extractions for Excel files (--columns is for CSV only)",
                }
                for col in columns
            ]
        else:
            results = [{"error": f"Unsupported file type: {suffix}"}]

    else:
        json.dump(
            [{"error": "Provide --extractions (Excel) or --columns (CSV)"}],
            sys.stdout,
            indent=2,
        )
        sys.exit(1)

    json.dump(results, sys.stdout, indent=2)
    print()  # trailing newline


if __name__ == "__main__":
    main()
