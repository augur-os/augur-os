"""ExcelAnalyzer: extract structure from .xlsx, .xls, and .csv files (ADR-086)."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

# openpyxl is an optional dependency
try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    HAS_OPENPYXL = False

# Patterns that suggest a totals/summary row
TOTALS_PATTERNS = re.compile(r"(?i)\b(total|sum|net|balance|grand\s*total|subtotal)\b")


class ExcelAnalyzer:
    """Analyze Excel and CSV files to extract structure.

    Returns a FileStructure dict with:
        - type: "excel" | "csv"
        - sheets: list of sheet info dicts (xlsx only)
        - columns: list of column header names
        - row_count: total data rows
        - has_totals: whether a totals row was detected
        - has_formulas: whether formulas were found (xlsx only)
        - totals_cells: list of {sheet, cell, label} for detected totals
    """

    def analyze(self, path: Path) -> dict[str, Any]:
        """Analyze an Excel or CSV file.

        Args:
            path: Path to the file.

        Returns:
            FileStructure dict.
        """
        path = Path(path)
        suffix = path.suffix.lower()

        if suffix == ".csv":
            return self._analyze_csv(path)
        elif suffix in (".xlsx", ".xls"):
            return self._analyze_xlsx(path)

        return {"type": "unknown", "error": f"Unsupported extension: {suffix}"}

    def _analyze_csv(self, path: Path) -> dict[str, Any]:
        """Analyze a CSV file."""
        result: dict[str, Any] = {
            "type": "csv",
            "columns": [],
            "row_count": 0,
            "has_totals": False,
            "totals_cells": [],
        }

        try:
            with open(path, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                if headers:
                    result["columns"] = [h.strip() for h in headers if h.strip()]

                row_count = 0
                last_row: list[str] = []
                for row in reader:
                    row_count += 1
                    if row:
                        last_row = row

                result["row_count"] = row_count

                # Check if last row looks like totals
                if last_row:
                    first_cell = last_row[0] if last_row else ""
                    if TOTALS_PATTERNS.search(first_cell):
                        result["has_totals"] = True
        except Exception as e:
            result["error"] = str(e)

        return result

    def _analyze_xlsx(self, path: Path) -> dict[str, Any]:
        """Analyze an Excel .xlsx file using openpyxl."""
        result: dict[str, Any] = {
            "type": "excel",
            "sheets": [],
            "columns": [],
            "row_count": 0,
            "has_totals": False,
            "has_formulas": False,
            "totals_cells": [],
        }

        if not HAS_OPENPYXL:
            result["error"] = "openpyxl not installed — install for full Excel analysis"
            # Fall back to basic metadata
            result["type"] = "excel_no_lib"
            return result

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=False)
            sheets_info = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data: dict[str, Any] = {
                    "name": sheet_name,
                    "columns": [],
                    "row_count": 0,
                    "has_totals": False,
                    "has_formulas": False,
                    "totals_cells": [],
                }

                rows = list(ws.iter_rows(max_row=min(ws.max_row or 1, 500)))
                if not rows:
                    sheets_info.append(sheet_data)
                    continue

                # Extract column headers from first row
                header_row = rows[0]
                headers = []
                for cell in header_row:
                    val = cell.value
                    if val is not None:
                        headers.append(str(val).strip())
                sheet_data["columns"] = headers

                # Use first sheet's columns as the top-level columns
                if not result["columns"] and headers:
                    result["columns"] = headers

                sheet_data["row_count"] = len(rows) - 1  # subtract header

                # Scan for formulas and totals
                for row in rows:
                    for cell in row:
                        # Check for formulas
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            sheet_data["has_formulas"] = True
                            result["has_formulas"] = True

                        # Check for totals labels
                        if cell.value and isinstance(cell.value, str):
                            if TOTALS_PATTERNS.search(cell.value):
                                sheet_data["has_totals"] = True
                                result["has_totals"] = True
                                # Record the cell reference for the value next to the label
                                if cell.column is not None and cell.row is not None:
                                    col_letter = openpyxl.utils.get_column_letter(cell.column + 1)
                                    sheet_data["totals_cells"].append(
                                        {
                                            "sheet": sheet_name,
                                            "cell": f"{col_letter}{cell.row}",
                                            "label": str(cell.value).strip(),
                                        }
                                    )

                sheets_info.append(sheet_data)

            result["sheets"] = sheets_info
            result["row_count"] = sum(s["row_count"] for s in sheets_info)
            result["totals_cells"] = [tc for s in sheets_info for tc in s["totals_cells"]]

            wb.close()
        except Exception as e:
            result["error"] = str(e)

        return result
